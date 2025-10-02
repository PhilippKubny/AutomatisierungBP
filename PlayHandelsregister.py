#!/usr/bin/env python3
"""
Playwright-based CLI for handelsregister.de

- Performs "Advanced search"
- Lists result rows (court, name, state, status)
- Optional: downloads "AD" (Current hard copy printout) PDFs for each result
  to the given --outdir using the filename: <Firmenname>_<YYYY-MM-DD>.pdf

Note: Respects site behavior; do not exceed rate limits.
"""


import argparse
import asyncio
import os
import re
import time
import sys
from datetime import date
from typing import Optional, Callable, Iterable

from openpyxl import load_workbook

import exel
import PDFScanner

from playwright.async_api import async_playwright, TimeoutError as PWTimeoutError, Locator


page = None  # Global page object for async context
counter = 0 #for timer logic
reruns = 0 #for reruns logic
debug = False
debug_scan = False


LocatorFactory = Callable[[], Locator]


async def _first_existing_locator(candidates: Iterable[LocatorFactory]) -> Optional[Locator]:
    """Return the first locator from *candidates* that resolves to at least one element."""
    for make_locator in candidates:
        try:
            locator = make_locator()
        except Exception:
            continue
        try:
            if await locator.count() > 0:
                return locator
        except Exception:
            continue
    return None


async def _fill_candidate(
    page,
    candidates: Iterable[LocatorFactory],
    value: str,
    label: str,
    *,
    debug_enabled: bool = False,
):
    """Try to fill *value* into the first matching locator from *candidates*.

    Returns the locator on success so callers can reuse it (e.g. to press Enter), otherwise ``None``.
    """

    locator = await _first_existing_locator(candidates)
    if locator is None:
        if debug_enabled:
            print(f"[{label}] no candidate locator matched.")
        return None

    try:
        await locator.fill(value, timeout=4000)
        if debug_enabled:
            print(f"[{label}] filled using locator {await locator.evaluate('el => el.tagName')}")
        return locator
    except Exception:
        # Fallback: try to manually clear via keyboard before typing
        try:
            await locator.click(timeout=2000)
            await locator.press("Control+A")
            await locator.type(value, delay=20)
            if debug_enabled:
                print(f"[{label}] typed value via keyboard fallback.")
            return locator
        except Exception as exc:
            if debug_enabled:
                print(f"[{label}] failed to fill value: {exc}")
            return None


async def _check_candidate(
    page,
    candidates: Iterable[LocatorFactory],
    label: str,
    *,
    debug_enabled: bool = False,
) -> bool:
    """Check (or click) the first radio/checkbox candidate."""

    locator = await _first_existing_locator(candidates)
    if locator is None:
        if debug_enabled:
            print(f"[{label}] no candidate locator matched.")
        return False

    try:
        await locator.check(timeout=4000)
        if debug_enabled:
            print(f"[{label}] checked via locator.")
        return True
    except Exception:
        try:
            await locator.click(timeout=4000)
            if debug_enabled:
                print(f"[{label}] clicked locator as fallback.")
            return True
        except Exception as exc:
            if debug_enabled:
                print(f"[{label}] failed to activate locator: {exc}")
            return False


async def _click_candidate(
    page,
    candidates: Iterable[LocatorFactory],
    label: str,
    *,
    debug_enabled: bool = False,
    delay_ms: int = 250,
) -> bool:
    """Click the first matching locator from *candidates*."""

    locator = await _first_existing_locator(candidates)
    if locator is None:
        if debug_enabled:
            print(f"[{label}] no candidate locator matched.")
        return False

    try:
        await locator.click(timeout=4000)
        if delay_ms:
            await page.wait_for_timeout(delay_ms)
        if debug_enabled:
            text = ""
            try:
                text = (await locator.inner_text()).strip()
            except Exception:
                pass
            print(f"[{label}] clicked locator with text '{text}'.")
        return True
    except Exception as exc:
        if debug_enabled:
            print(f"[{label}] failed to click locator: {exc}")
        return False


async def _wait_for_any_selector(page, selectors: Iterable[str], timeout: int = 30000) -> Optional[str]:
    """Wait until any selector from *selectors* appears and return the one that matched."""

    deadline = time.monotonic() + (timeout / 1000.0)
    last_error: Optional[Exception] = None
    for selector in selectors:
        remaining = int((deadline - time.monotonic()) * 1000)
        if remaining <= 0:
            break
        try:
            await page.wait_for_selector(selector, timeout=remaining)
            return selector
        except Exception as exc:
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    raise PWTimeoutError(f"Timeout waiting for selectors: {', '.join(selectors)}")


async def open_advanced(page, debug: bool = False) -> bool:
    candidates = [
        lambda: page.get_by_role("button", name=re.compile(r"(Erweiterte|Weitere)\\s+Suche", re.I)).first(),
        lambda: page.get_by_role("link", name=re.compile(r"(Erweiterte|Weitere)\\s+Suche", re.I)).first(),
        lambda: page.locator("text=/Erweiterte\\s*Suche/i").first(),
        lambda: page.locator("text=/Weitere\\s*Suchoptionen/i").first(),
        lambda: page.locator("text=/Weitere\\s*Suchkriterien/i").first(),
        lambda: page.locator("button:has-text('Erweiterte Suche')").first(),
        lambda: page.locator("a:has-text('Erweiterte Suche')").first(),
        lambda: page.locator("[aria-controls*='erweitert' i]").first(),
        lambda: page.locator("[data-action*='advanced' i], [data-testid*='advanced' i]").first(),
        lambda: page.locator("[id*='erweitert' i], [name*='erweitert' i]").first(),
    ]
    clicked = await _click_candidate(
        page,
        candidates,
        "advanced",
        debug_enabled=debug,
        delay_ms=350,
    )
    if clicked:
        return True

    try:
        await page.get_by_label(re.compile(r"(Postleitzahl|PLZ)", re.I)).wait_for(timeout=1500)
        if debug:
            print("[advanced] form appears open (label found)")
        return True
    except Exception:
        pass

    if debug:
        print("[advanced] not found via all strategies")
    return False


async def set_plz(page, plz: str, debug: bool = False) -> bool:
    if not plz:
        return False
    candidates = [
        lambda: page.get_by_label(re.compile(r"(Postleitzahl|PLZ)", re.I)),
        lambda: page.get_by_placeholder(re.compile(r"(Postleitzahl|PLZ|ZIP|Postcode)", re.I)),
        lambda: page.get_by_role("textbox", name=re.compile(r"(Postleitzahl|PLZ)", re.I)).first(),
        lambda: page.locator("[aria-label*='PLZ' i], [aria-label*='Postleit' i], [aria-label*='zip' i]").first(),
        lambda: page.locator("input[id*='plz' i], input[name*='plz' i]").first(),
        lambda: page.locator("input[id*='postleit' i], input[name*='postleit' i]").first(),
        lambda: page.locator("input[id*='zip' i], input[name*='zip' i]").first(),
        lambda: page.locator("input[title*='PLZ' i], input[data-label*='PLZ' i]").first(),
    ]
    locator = await _fill_candidate(
        page,
        candidates,
        plz,
        "plz",
        debug_enabled=debug,
    )
    if locator:
        return True
    if debug:
        print("[plz] no candidate matched")
    return False


async def scan_ui(page):
    print("\n--- UI SCAN: BUTTONS/LINKS ---")
    for sel in ["button, [role=button], a"]:
        loc = page.locator(sel)
        n = await loc.count()
        for i in range(n):
            el = loc.nth(i)
            try:
                txt = (await el.inner_text()).strip()
            except Exception:
                txt = ""
            try:
                rid = await el.get_attribute("id")
                name = await el.get_attribute("name")
                arial = await el.get_attribute("aria-label")
                role = await el.get_attribute("role")
            except Exception:
                rid = name = arial = role = None
            print(f"[btn] text='{txt}' id='{rid}' name='{name}' aria-label='{arial}' role='{role}'")

    print("\n--- UI SCAN: INPUTS ---")
    inputs = page.locator("input, select, textarea")
    m = await inputs.count()
    for i in range(m):
        el = inputs.nth(i)
        try:
            rid = await el.get_attribute("id")
            name = await el.get_attribute("name")
            ph = await el.get_attribute("placeholder")
            arial = await el.get_attribute("aria-label")
            typ = await el.get_attribute("type")
        except Exception:
            rid = name = ph = arial = typ = None
        label_txt = ""
        if rid:
            label = page.locator(f"label[for='{rid}']")
            if await label.count() > 0:
                try:
                    label_txt = (await label.first().inner_text()).strip()
                except Exception:
                    pass
        print(f"[in] id='{rid}' name='{name}' type='{typ}' placeholder='{ph}' aria-label='{arial}' label='{label_txt}'")
    print("--- UI SCAN END ---\n")


async def goto_with_retry(page, url: str, attempts: int = 3, base_timeout_ms: int = 60000):
    last_err = None
    for i in range(1, attempts + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=base_timeout_ms)
            try:
                await page.get_by_role(
                    "button",
                    name=re.compile(r"(Alle\s+akzeptieren|Akzeptieren|Zustimmen)", re.I),
                ).click(timeout=3000)
            except Exception:
                pass
            try:
                await _wait_for_any_selector(
                    page,
                    [
                        "text=Erweiterte Suche",
                        "text=Weitere Suchoptionen",
                        "text=Weitere Suchkriterien",
                        "text=Erweiterte Suchkriterien",
                        "text=Erweiterte Suche anzeigen",
                        "input[type='text']",
                    ],
                    timeout=12000,
                )
            except Exception:
                await page.wait_for_selector("input[type='text']", timeout=5000)
            return
        except Exception as e:
            last_err = e
            await asyncio.sleep(1.5 * i)
    raise last_err


def apply_page_timeouts(target_page) -> None:
    target_page.set_default_navigation_timeout(90000)
    target_page.set_default_timeout(45000)

# Map to existing CLI semantics
SCHLAGWORT_OPTIONEN = {
    "all": 1,    # contain all keywords
    "min": 2,    # contain at least one keyword
    "exact": 3,  # exact company name
}


def read_cell(ws, row: int, col_letter: str) -> Optional[str]:
    """Return the trimmed value of a worksheet cell or ``None`` if empty."""
    if ws is None or not col_letter:
        return None
    cell_ref = f"{col_letter}{row}"
    try:
        value = ws[cell_ref].value
    except Exception as exc:  # pragma: no cover - defensive logging
        print(f"[warn] Could not read cell {cell_ref}: {exc}")
        return None
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def looks_like_same_zip(text: Optional[str], plz: Optional[str]) -> bool:
    """Return ``True`` if *text* contains the exact *plz* as a standalone token."""
    if not text or not plz:
        return False
    pattern = re.compile(rf"\b{re.escape(plz)}\b")
    return bool(pattern.search(text))


# TODO: Debug helper to print element value and outerHTML
async def _debug_dump_element(page, selector: str, label: str, clip: int = 1500):
    """Print value + outerHTML of an element (trimmed), only for debug."""
    try:
        el = page.locator(selector).first
        # Works for <input> and <textarea>
        try:
            current_value = await el.input_value()
        except Exception:
            current_value = "<no input_value()>"
        outer = await el.evaluate("n => n.outerHTML")
        short = outer if len(outer) <= clip else (outer[:clip] + "…[truncated]")
        print(f"[debug] {label} value:", repr(current_value))
        print(f"[debug] {label} outerHTML:", short)
    except Exception as e:
        print(f"[debug] Could not dump {label}: {e}")

async def _debug_dump_results(page, clip: int = 5000):
    """
    Dump outerHTML of the results area after search.
    Tries ergebnisForm first, then its result table, then the grid.
    """
    selectors = [
        "form#ergebnisForm",
        "form[id^='ergebnisForm']",
        "#ergebnisForm\\:selectedSuchErgebnisFormTable_data",
        "[id$='selectedSuchErgebnisFormTable_data']",
        "table[role='grid']",
    ]
    for sel in selectors:
        try:
            await page.wait_for_selector(sel, timeout=6000)
            el = page.locator(sel).first
            html = await el.evaluate("n => n.outerHTML")
            short = html if len(html) <= clip else (html[:clip] + "…[truncated]")
            print(f"[debug] results HTML from {sel}:\n{short}")
            return
        except Exception:
            continue

    # Last resort: dump body
    try:
        body_html = await page.locator("body").evaluate("n => n.outerHTML")
        short = body_html if len(body_html) <= clip else (body_html[:clip] + "…[truncated]")
        print("[debug] results section not found; dumping <body> instead:\n", short)
    except Exception as e:
        print(f"[debug] could not dump body: {e}")

#TODO: No reach, errors

async def rerun_search(keyword: str, mode: str, register_number: str = None, postal_code: str = None, download=False, company_name=None, sap_number=None, outdir=None):
    global page
    global reruns
    reruns += 1
    if reruns > 3:
        # sleep for 10 minutes if more than 3 reruns
        print("[warn] More than 3 reruns, sleeping for 10 minutes to avoid rate limiting...")
        time.sleep(600)
        reruns = 0
    print("[warn] Rerun")
    page = await page.context.new_page()  # Reset page context
    apply_page_timeouts(page)
    await open_startpage()
    await perform_search(keyword, mode, register_number=register_number, postal_code=postal_code)
    if download:
        return await download_ad_for_row(company_name, outdir, sap_number)
    return None



#TODO: Perform search

async def open_startpage():
    # Landing page → ensure we land on welcome.xhtml
    global page
    global debug
    await goto_with_retry(
        page,
        "https://www.handelsregister.de/rp_web/welcome.xhtml",
        attempts=4,
        base_timeout_ms=90000,
    )
    if debug:
        print("[debug] opened welcome page:", page.url)


async def perform_search(keyword: str, mode: str, register_number: str = None, postal_code: str = None):
    """
    Click 'Advanced search', fill the form, submit.
    """
    global page
    global counter
    global debug
    global debug_scan

    opened = await open_advanced(page, debug=debug_scan)
    if not opened and debug_scan:
        await scan_ui(page)

    try:
        await _wait_for_any_selector(
            page,
            [
                "#form\\:schlagwoerter",
                "[id$='schlagwoerter']",
                "[name$='schlagwoerter']",
                "input[id*='schlagwort' i]",
                "input[name*='schlagwort' i]",
                "input[placeholder*='Schlagwort' i]",
                "input[aria-label*='Schlagwort' i]",
                "input[id*='suchbegriff' i]",
                "input[name*='suchbegriff' i]",
                "input[placeholder*='Suchbegriff' i]",
                "input[aria-label*='Suchbegriff' i]",
                "input[id*='firma' i]",
                "input[name*='firma' i]",
                "input[placeholder*='Firma' i]",
                "input[aria-label*='Firma' i]",
                "text=/Schlagwort/i",
                "text=/Suchbegriff/i",
                "text=/Firma/i",
            ],
            timeout=30000,
        )
    except Exception:
        print("[warn] Advanced search form not found; Website not reachable or UI may have changed.")
        await rerun_search(keyword, mode, register_number, postal_code)
        return

    words = [w for w in re.split(r"[ -]+", keyword) if w]
    keyword = " ".join(words[:5])  # Limit to first 5 words for search

    keyword_locator = await _fill_candidate(
        page,
        [
            lambda: page.locator("#form\\:schlagwoerter"),
            lambda: page.locator("[id$='schlagwoerter']").first(),
            lambda: page.locator("[name$='schlagwoerter']").first(),
            lambda: page.get_by_label(re.compile(r"Schlagwort", re.I)).first(),
            lambda: page.get_by_placeholder(re.compile(r"Schlagwort", re.I)).first(),
            lambda: page.get_by_role("textbox", name=re.compile(r"Schlagwort", re.I)).first(),
            lambda: page.locator("[id*='suchbegriff' i]").first(),
            lambda: page.locator("[name*='suchbegriff' i]").first(),
            lambda: page.get_by_label(re.compile(r"Suchbegriff", re.I)).first(),
            lambda: page.get_by_placeholder(re.compile(r"Suchbegriff", re.I)).first(),
            lambda: page.get_by_role("textbox", name=re.compile(r"Suchbegriff", re.I)).first(),
            lambda: page.locator("[id*='firma' i]").first(),
            lambda: page.locator("[name*='firma' i]").first(),
            lambda: page.get_by_label(re.compile(r"Firma|Unternehmensname", re.I)).first(),
            lambda: page.get_by_placeholder(re.compile(r"Firma|Unternehmensname", re.I)).first(),
            lambda: page.get_by_role("textbox", name=re.compile(r"Firma|Unternehmensname", re.I)).first(),
        ],
        keyword,
        "schlagwoerter",
        debug_enabled=debug_scan,
    )
    if keyword_locator is None:
        print("[warn] Could not fill keyword field; Website not reachable or UI may have changed.")
        await rerun_search(keyword, mode, register_number, postal_code)
        return

    register_number = (register_number or "").strip()
    register_locator = await _fill_candidate(
        page,
        [
            lambda: page.locator("#form\\:registerNummer"),
            lambda: page.locator("[id$='registerNummer']").first(),
            lambda: page.locator("[name$='registerNummer']").first(),
            lambda: page.get_by_label(re.compile(r"Register-?nummer", re.I)).first(),
            lambda: page.get_by_placeholder(re.compile(r"Register-?nummer", re.I)).first(),
            lambda: page.get_by_role("textbox", name=re.compile(r"Register", re.I)).first(),
        ],
        register_number,
        "registernummer",
        debug_enabled=debug_scan,
    )
    if register_locator is None and register_number:
        print("[warn] Could not fill register number field; Website not reachable or UI may have changed.")
        await rerun_search(keyword, mode, register_number, postal_code)
        return

    postal_code_value = str(postal_code).strip() if postal_code else ""
    if postal_code_value:
        ok = await set_plz(page, postal_code_value, debug=debug_scan)
        if ok:
            print(f"[info] Postal code '{postal_code_value}' set.")
        else:
            print(f"[warn] Could not set postal code '{postal_code_value}'.")
            if debug_scan:
                print("[hint] PLZ-Selektoren passen nicht. Siehe UI-SCAN oben und ergänze eine passende Heuristik.")
    else:
        if debug:
            print(f"[debug] Postal code filter inactive for '{keyword}'.")


    # Radio/select for schlagwortOptionen:
    so_value = SCHLAGWORT_OPTIONEN[mode]
    radio_candidates = [
        lambda: page.locator(f"input[name='form:schlagwortOptionen'][value='{so_value}']"),
        lambda: page.locator(f"input[value='{so_value}'][name*='schlagwortOptionen']"),
        lambda: page.locator(f"[id*='schlagwortOptionen'] input[value='{so_value}']").first(),
        lambda: page.locator(f"input[type='radio'][value='{so_value}']").first(),
    ]
    if not await _check_candidate(page, radio_candidates, "schlagwortOptionen", debug_enabled=debug_scan):
        print("[warn] Could not activate keyword mode radio button; Website not reachable or UI may have changed.")
        await rerun_search(keyword, mode, register_number, postal_code)
        return

    search_candidates = [
        lambda: page.locator("#form\\:btnSuche"),
        lambda: page.locator("[id$='btnSuche']").first(),
        lambda: page.locator("[id*='suche'][type='submit']").first(),
        lambda: page.locator("input[type='submit'][value*='Suche' i]").first(),
        lambda: page.get_by_role("button", name=re.compile(r"Suche\\s*(starten|ausführen|beginnen)?", re.I)).first(),
        lambda: page.get_by_role("button", name=re.compile(r"Suchen", re.I)).first(),
        lambda: page.locator("button:has-text('Suche')").first(),
    ]
    clicked = await _click_candidate(
        page,
        search_candidates,
        "search-button",
        debug_enabled=debug_scan,
        delay_ms=400,
    )

    if not clicked:
        pressed_enter = False
        try:
            if keyword_locator is not None:
                await keyword_locator.press("Enter")
                pressed_enter = True
                if debug_scan:
                    print("[search-button] pressed Enter in keyword field as fallback.")
        except Exception:
            pressed_enter = False

        if not pressed_enter:
            print("[warn] Could not trigger search; Website not reachable or UI may have changed.")
            await rerun_search(keyword, mode, register_number, postal_code)
            return


    if debug:
        print("[debug] clicked search; waiting for results…")

    counter += 1 # Successfully clicked on search -> counter +1
    # Wait for results table, check for specific section in HTML body
    try:
        await _wait_for_any_selector(
            page,
            [
                "#ergebnissForm\\:selectedSuchErgebnisFormTable_data",
                "#ergebnisForm\\:selectedSuchErgebnisFormTable_data",
                "[id$='selectedSuchErgebnisFormTable_data']",
                "table[role='grid']",
                "table[id*='Ergebnis']",
                "text=/Keine Daten|Kein Ergebnis/i",
            ],
            timeout=30000,
        )
    except Exception:
        print("[warn] Results table not found; Website not reachable or UI may have changed.")
        await rerun_search(keyword, mode, register_number, postal_code)
        return

    if debug:
        #await _debug_dump_results(page)
        print("[debug] results page loaded!")

async def get_results(postal_code: Optional[str] = None):
    """
    Scrape the visible rows of the results table (first page).
    Returns list of dicts with minimal fields, and row locators for clicking AD per row.
    """
    global debug
    row_selector = (
        "table[role='grid'] tr[data-ri], "
        "table[id*='Ergebnis'] tr[data-ri], "
        "table[id*='Ergebnis'] tbody tr"
    )
    rows = page.locator("".join(row_selector))
    count = await rows.count()
    results = []

    for i in range(count):
        row = rows.nth(i)

        try:
            if not await row.is_visible():
                continue
        except Exception:
            pass

        cells = row.locator("td")
        ccount = await cells.count()
        if ccount == 0:
            continue

        texts = []
        for j in range(ccount):
            try:
                t = (await cells.nth(j).inner_text()).strip()
            except Exception:
                t = ""
            texts.append(t)

        row_text = " ".join(texts).strip()
        if not row_text:
            continue
        if re.search(r"Keine\s+(Treffer|Daten|Ergebnisse)", row_text, re.I):
            if debug:
                print(f"[debug] Skipping non-result row {i}: {row_text}")
            continue

        def _safe(idx: int, fallback: str = "") -> str:
            if len(texts) > idx and texts[idx]:
                return texts[idx]
            return fallback

        court = _safe(1, _safe(0, ""))
        name = _safe(2, _safe(1, ""))
        registered_office = _safe(3, _safe(2, ""))
        status = _safe(4, _safe(3, ""))

        address = None
        address_source = None
        for selector in (
            ".address",
            ".ergebnisAdresse",
            ".result-address",
            "[data-label*='Adresse']",
            "td[data-label*='Ort']",
        ):
            try:
                locator = row.locator(selector)
                if await locator.count():
                    address = (await locator.first.inner_text()).strip()
                    address_source = selector
                    break
            except Exception:
                continue

        address = address or registered_office
        if postal_code and debug and address_source:
            print(f"[debug] Result row {i} address from {address_source}: {address}")

        results.append(
            {
                "row_index": i,
                "court": court,
                "name": name,
                "registered_office": registered_office,
                "status": status,
                "address": address,
                "row_locator": row,  # keep for AD click
            }
        )

    if postal_code:
        filtered = [r for r in results if looks_like_same_zip(r.get("address"), postal_code)]
        dropped = len(results) - len(filtered)
        if dropped:
            print(f"[info] Postal code '{postal_code}' filtered out {dropped} result(s).")
        print(f"[info] Postal code '{postal_code}' left {len(filtered)} of {len(results)} result(s).")
        return filtered

    return results


#TODO: Download functions

def sanitize_filename(name: str) -> str:
    # Windows-safe filename
    name = name.strip()
    return re.sub(r'[\\/:*?"<>|]+', "_", name)

def create_human_check_file():
    """Create a HumanCheck text file in Downloads directory for searches with multiple results or none"""
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    check_file = os.path.join(downloads_path, "HumanCheck.txt")

    if os.path.exists(check_file):
        return

    try:
        with open(check_file, "w") as f:
            f.write(f"Script execution started at: {date.today().strftime('%d-%m-%Y %H:%M:%S')}")
        print(f"[info] Created human check file: {check_file}")
    except Exception as e:
        print(f"[error] Failed to create human check file: {e}")

def replace_umlauts(text):
    """Replace German umlauts after uppercase conversion"""
    replacements = {
        'Ö': 'OE',
        'Ä': 'AE',
        'Ü': 'UE',
        'ß': 'SS'
    }
    for umlaut, replacement in replacements.items():
        text = text.replace(umlaut, replacement)
    return text

async def download_ad_for_row(company_name, outdir, sap_number=None, row_locator=None):
    """
    Waits for the single search result, clicks the AD link, and saves the PDF as
    '<Company>_dd.mm.yyyy.pdf'. Returns the saved path or None on failure.
    """
    #for now only one row is expected,
    #TODO: if more than one row is found, iterate over them, use row_locator
    #Uppercase and replace umlauts in company name
    global page
    global debug
    try:
        #Check if the outdir exists, if not create it
        os.makedirs(outdir, exist_ok=True)

        # Most specific & stable selector for AD (per your HTML)
        # Trigger and capture the download
        ad_candidates = []
        if row_locator is not None:
            ad_candidates.extend(
                [
                    lambda: row_locator.locator("a[onclick*='Dokumentart.AD']").first(),
                    lambda: row_locator.get_by_role("link", name=re.compile(r"\bAD\b", re.I)).first(),
                    lambda: row_locator.locator("button:has-text('AD')").first(),
                ]
            )
        ad_candidates.extend(
            [
                lambda: page.locator("#ergebnissForm\\:selectedSuchErgebnisFormTable_data a[onclick*='Dokumentart.AD']").first(),
                lambda: page.locator("#ergebnisForm\\:selectedSuchErgebnisFormTable_data a[onclick*='Dokumentart.AD']").first(),
                lambda: page.locator("[id$='selectedSuchErgebnisFormTable_data'] a[onclick*='Dokumentart.AD']").first(),
                lambda: page.locator("a[data-dokumentart='AD']").first(),
                lambda: page.get_by_role("link", name=re.compile(r"\bAD\b", re.I)).first(),
                lambda: page.locator("a:has-text('AD')").first(),
            ]
        )

        ad_locator = await _first_existing_locator(ad_candidates)
        if ad_locator is None:
            print(f"[warn] Failed to locate AD link for '{company_name}'; download may not have started.")
            return None

        try:
            await ad_locator.scroll_into_view_if_needed()
        except Exception:
            pass

        try:
            async with page.expect_download(timeout=40000) as dl_info:
                await ad_locator.click()
            download = await dl_info.value
            if debug:
                print(f"[debug] Download started for '{company_name}': {download.suggested_filename}")
        except Exception:
            print(f"[warn] Failed to click AD link for '{company_name}'; download may not have started.")
            return None

        # sap_company_dd.mm.yyyy filename
        date_str = date.today().strftime("%d-%m-%Y")
        prefix = (sanitize_filename(sap_number) + "_") if sap_number else ""
        fname = f"{prefix}{sanitize_filename(company_name)}_{date_str}.pdf"
        save_path = os.path.join(outdir, fname)

        await download.save_as(save_path)
        # Verify the file exists and has size > 0 (not corrupted)
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            if debug:
                print(f"[debug] Saved AD PDF: {save_path}")
            return save_path

    except Exception as e:
        if debug:
            print(f"[warn] Failed to download AD for '{company_name}': {e}")
        return None


#TODO: Main

async def main_async(args):
    """
        Main asynchronous entry point for the Handelsregister script.
        Depending on CLI args, runs either:
          - Excel batch mode: read multiple company names/register numbers from Excel
          - Single-shot mode: run for a single provided search term

        Handles:
          - Opening the Handelsregister start page
          - Performing search queries
          - Downloading 'AD' PDF documents if requested
          - Logging cases where results are ambiguous
    """
    global page
    global counter
    global debug
    global debug_scan

    debug = args.debug
    debug_scan = getattr(args, "debug_scan", False)
    # Ensure output directory exists
    if not os.path.exists(args.outdir):
        default_path = os.path.join(os.path.expanduser("~"), "Downloads", "BP")
        print(f"[warn] Path not found: {args.outdir}")
        print(f"[info] Creating default directory: {default_path}")
        os.makedirs(default_path, exist_ok=True)
        args.outdir = default_path
    else:
        os.makedirs(args.outdir, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=not args.headful, slow_mo=0)
        context = await browser.new_context(
            accept_downloads=True,
            locale="de-DE",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
        )

        await context.route(
            re.compile(r".*\.(png|jpe?g|gif|svg|webp)(\?.*)?$", re.I),
            lambda request: request.abort(),
        )

        page = await context.new_page()
        apply_page_timeouts(page)

        if debug:
            print("[debug] Browser launched, opening start page...")
        await open_startpage()




        #TODO: Excel batch mode
        if args.excel:
            # Read company/job data from Excel using helper function
            jobs = exel.read_jobs_from_excel(
                path=args.excel,
                sheet=args.sheet,
                name_col=args.name_col,
                regno_col=args.regno_col,
                sap_supplier_col=args.sap_supplier_col,
                sap_customer_col=args.sap_customer_col,
                postal_code_col=args.postal_code_col,
                country_col=args.country_col,
                start=args.start,
                end=args.end,
            )

            print(f"[info] loaded {len(jobs)} jobs from Excel (rows {args.start or 3}..{args.end or 'last'})")

            ws_postal = None
            wb_postal = None
            if args.postal_code:
                try:
                    wb_postal = load_workbook(args.excel, data_only=True, read_only=True)
                    ws_postal = wb_postal[args.sheet] if args.sheet else wb_postal.active
                except Exception as exc:
                    print(f"[warn] Could not open workbook for postal code lookup: {exc}")

            start_time = time.time() # 1 hour timer to not go over the 60 search limitation (VERY IMPORTANT)
            # Iterate through each job (company) from the Excel list
            for i, job in enumerate(jobs, 1):
                # TODO: Other countries
                if job["country"] != "DE":
                    print(f"[warn] Skipping job {i}: country is not DE (got '{job['country']}').")
                    continue

                # Timer logic: every 60 jobs, wait for the remaining time to complete the hour
                if counter >= 60:
                    elapsed_time = time.time() - start_time
                    hour_in_seconds = 3600

                    # If less than an hour has passed, wait for remaining time
                    if elapsed_time < hour_in_seconds:
                        wait_time = hour_in_seconds - elapsed_time
                        print(f"[info] Waiting {wait_time:.0f} seconds to complete the hour...")
                        await page.wait_for_timeout(wait_time * 1000)  # convert to milliseconds

                    # Reset the timer
                    start_time = time.time()
                    print("[info] Hour timer reset")
                    page = await context.new_page()  # Reset page context
                    apply_page_timeouts(page)
                    await open_startpage()


                if job["name"] is None:
                    print(f"[warn] Skipping job {i}: no company name provided.")

                    continue

                kw = job["name"]  # Company name to search for
                reg = str(job["register_no"]) if job["register_no"] is not None else ""  # Register number (if available), normalized if it was in float
                sap = job["sap"]  # SAP number (if available)
                row_number = i + args.start - 1
                postal_code = None
                if args.postal_code:
                    candidate = job.get("postal_code") if isinstance(job, dict) else None
                    postal_code = str(candidate).strip() if candidate else None
                    if not postal_code and ws_postal is not None:
                        postal_code = read_cell(ws_postal, row_number, args.postal_code_col)
                    if postal_code:
                        print(f"[info] Using postal code '{postal_code}' for row {row_number}.")
                    else:
                        print(f"[info] No postal code for row {row_number}; running search without PLZ filter.")

                if debug:
                    print("")
                    print(f"[debug] ({i}/{len(jobs)}) {sap or 'NoSAP'} | {kw} | reg={reg or 'None'}")

                # Refresh the start page for each job (avoids leftover form state) TODO: check if this is needed
                #await open_startpage(page, debug=debug)

                # Perform the advanced search with the name + optional register number
                await perform_search(kw, args.schlagwortOptionen, register_number=reg, postal_code=postal_code)
                print(f"[debug] {kw} | reg={reg or 'None'}")

                # Retrieve the search results (list of rows)
                results = await get_results(postal_code if args.postal_code else None)

                # If we don't have exactly one match, log it to HumanCheck.txt and skip
                check_file = os.path.join(os.path.expanduser("~"), "Downloads", "HumanCheck.txt")
                f = open(check_file, "a")
                if len(results) != 1:
                    f.write(f"\n[info] found {len(results)} result row(s) for '{kw}' (SAP={sap or 'None'})")
                    print(f"[warn] {kw}: expected 1 result, got {len(results)} → logged to HumanCheck.txt")
                    exel.write_to_excel_error(
                        path=args.excel,
                        sheet=args.sheet,
                        row=i + args.start - 1,  # Adjust for 0-based index
                        changes_check_col=args.changes_check_col,
                        error_col=args.name1_col,
                        error_msg=f"{len(results)}"
                    )
                    print(
                        f"[warn] Failed '{kw}' (SAP={sap or 'None'}); marked row {i + args.start - 1} in Excel as error.")
                    continue  # Skip to next company

                # If PDF download is enabled, download the AD (Current hard copy printout)
                r = results[0]  # The single matching result TODO: Multiple if needed
                company_name = replace_umlauts(r["name"].upper()) # Uppercase and replace umlauts
                if args.download_ad:
                    path = await download_ad_for_row(
                        company_name=company_name,
                        outdir=args.outdir,
                        sap_number=sap,  # Prefix SAP number to the filename if available
                        row_locator=r["row_locator"], #currently not used
                    )
                    while path is None:
                        path = await rerun_search(kw, args.schlagwortOptionen, register_number=reg, postal_code=postal_code, download=True, company_name=company_name, sap_number=sap, outdir=args.outdir)

                    update_info = PDFScanner.extract_from_pdf(path) | {"company_name": company_name, "sap_number": sap, "download_path": path} # Extract fields from the downloaded PDF into a dict, override company_name with umlauts replaced
                    if update_info["register_type"] == "unexpected Format":
                        f.write(f"[warn] Error, unexpected PDF Format '{company_name}' (SAP={sap or 'None'}) in row {i + args.start - 1}")
                        print(f"[warn] Error, unexpected PDF Format '{company_name}' (SAP={sap or 'None'})")
                        exel.write_to_excel_error(
                            path=args.excel,
                            sheet=args.sheet,
                            row=i + args.start - 1,  # Adjust for 0-based index
                            changes_check_col=args.changes_check_col,
                            error_col=args.name1_col,
                            error_msg=f"unexpected PDF Format",
                            pdf_path = path,  # Save the path of PDF in excel
                            pdf_path_col=args.doc_path_col,
                        )
                        continue
                    exel.write_update_to_excel(
                        path=args.excel,
                        sheet=args.sheet,
                        row=i + args.start - 1,  # Adjust for 0-based index
                        update_info=update_info, # All info to update
                        name_col=args.name1_col,
                        regno_col=args.regno_col,
                        sap_supplier_col=args.sap_supplier_col,
                        sap_customer_col=args.sap_customer_col,
                        name2_col=args.name2_col,
                        name3_col=args.name3_col,
                        street_col=args.street_col,
                        house_number_col=args.house_number_col,
                        city_col=args.city_col,
                        postal_code_col=args.postal_code_col,
                        doc_path_col=args.doc_path_col,
                        changes_check_col=args.changes_check_col,
                        date_check_col=args.date_check_col,
                        register_type_col=args.register_type_col,
                        check_file=os.path.join(os.path.expanduser("~"), "Downloads", "HumanCheck.txt")
                    )
                    print(f"[info] Updated Excel row {i + args.start - 1} for '{company_name}' (SAP={sap or 'None'})")

            if wb_postal is not None:
                wb_postal.close()


        #TODO: Single-shot mode
        else:
            if args.sap_number == "-1":
                print("In single-shot mode you must provide --sap_number.")
                return
            if not args.schlagwoerter:
                print("In single-shot mode you must provide --schlagwoerter.")
                return
            if args.row_number == "-1":
                print("In single-shot mode you must provide --row_number.")
                return
            single_postal_code = None
            if args.postal_code:
                single_postal_code = str(args.plz or "").strip() or None
                if single_postal_code:
                    print(f"[info] Using postal code '{single_postal_code}' for single-shot search.")
                else:
                    print("[info] Postal code flag enabled but no --plz value provided; continuing without filter.")
            await perform_search(args.schlagwoerter, args.schlagwortOptionen, register_number=args.register_number, postal_code=single_postal_code)
            results = await get_results(single_postal_code if args.postal_code else None)
            if len(results) != 1:
                # Write to HumanCheck.txt in Downloads
                check_file = os.path.join(os.path.expanduser("~"), "Downloads", "HumanCheck.txt")
                with open(check_file, "a") as f:
                    f.write(f"\n\n[info] found {len(results)} result row(s) for '{args.schlagwoerter}'")
                #    for r in results:
                #        f.write(f"\nname: {r['name']}")
                #        f.write(f"\ncourt: {r['court']}")
                #        f.write(f"\nstate/office: {r['registered_office']}")
                #        f.write(f"\nstatus: {r['status']}")
                #        f.write("\n")

                print("[warn] Found multiple results or none; check HumanCheck.txt in Downloads for details.")
                return
            #check if only one result is found, if more TODO: interation already exists

            path = None  # Initialize path to None
            if args.download_ad and len(results) == 1:
                # Process each row and click AD
                for r in results:
                    path = await download_ad_for_row(
                        company_name=r["name"] or "company",
                        outdir=args.outdir,
                        sap_number=args.sap_number, # No SAP number in this case TODO
                        row_locator=r["row_locator"], #currently not used
                    )
                    # Small pause to be gentle (adjust if needed)
                    #await page.wait_for_timeout(1200)
            
            if path is not None:
                    update_info = PDFScanner.extract_from_pdf(path) | {"company_name": results[0]["name"], "sap_number": args.sap_number, "download_path": path} # Extract fields from the downloaded PDF into a dict, override company_name with umlauts replaced
                    if update_info["register_type"] == "unexpected Format":
                        print(f"[warn] Error, unexpected PDF Format '{results[0]['name']}' (SAP={args.sap_number or 'None'})")
                        exel.write_to_excel_error(
                            path=args.excel,
                            sheet=args.sheet,
                            row=args.row_number,
                            changes_check_col=args.changes_check_col,
                            error_col=args.name1_col,
                            error_msg=f"unexpected PDF Format",
                            pdf_path = path,  # Save the path of PDF in excel
                            pdf_path_col=args.doc_path_col,
                        )
                        return
                    exel.write_update_to_excel(
                        path=args.excel,
                        sheet=args.sheet,
                        row=args.row_number,
                        update_info=update_info, # All info to update
                        name_col=args.name1_col, # Update column
                        regno_col=args.regno_col,
                        sap_supplier_col=args.sap_supplier_col,
                        sap_customer_col=args.sap_customer_col,
                        name2_col=args.name2_col,
                        name3_col=args.name3_col,
                        street_col=args.street_col,
                        house_number_col=args.house_number_col,
                        city_col=args.city_col,
                        postal_code_col=args.postal_code_col,
                        doc_path_col=args.doc_path_col,
                        changes_check_col=args.changes_check_col,
                        date_check_col=args.date_check_col,
                        register_type_col=args.register_type_col,
                        check_file=os.path.join(os.path.expanduser("~"), "Downloads", "HumanCheck.txt")
                    )
                    print(f"[info] Updated Excel row {args.row} for '{results[0]["name"]}' (SAP={args.sap_number or 'None'})")
        
                    

        await context.close()
        await browser.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="A handelsregister CLI (Playwright)")
    parser.add_argument(
        "-d", "--debug",
        default=False,
        help="Enable debug-style prints",
        action="store_true"
    )
    parser.add_argument(
        "--debug-scan",
        action="store_true",
        help="Dump UI (Buttons/Links/Inputs) zur Selektor-Diagnose",
    )
    parser.add_argument(
        "-s", "--schlagwoerter",
        help="Search for the provided keywords",
        default=None,
    )
    parser.add_argument(
        "-so", "--schlagwortOptionen",
        help="Keyword options: all=contain all keywords; min=contain at least one; exact=exact company name.",
        choices=["all", "min", "exact"],
        default="all"
    )
    parser.add_argument(
        "--download-ad",
        default=False,
        help="Download the AD (Current hard copy printout) PDF for each result row.",
        action="store_true",
    )
    parser.add_argument(
        "--outdir",
        help="Output directory for downloaded PDFs",
        default=os.path.join(os.path.expanduser("~"), "Downloads", "BP"),
    )
    parser.add_argument(
        "--headful",
        help="Run with a visible browser window (useful for debugging).",
        action="store_true",
    )
    parser.add_argument(
        "-rn", "--register-number",
        help="Optional: Handelsregisternummer to search for",
        required=False
    )
    parser.add_argument(
        "--postal-code",
        help="Enable postal code filtering for searches.",
        action="store_true",
    )
    parser.add_argument("--plz", default="", help="Postal code for single-shot searches (use with --postal-code).")
    parser.add_argument(
        "-sap", "--sap-number",
        default="-1",
        help="Optional: for singelshot only",
        required=False
    )
    parser.add_argument(
        "-row", "--row-number",
        default="-1",
        help="Optional: for singelshot only",
        required=False
    )
    # for excel import
    parser.add_argument("--excel", help="Path to Excel file (e.g., TestBP.xlsx)")
    parser.add_argument("--sheet", default=None, help="Optional sheet name")
    parser.add_argument("--name-col", default="C", help="Excel column with Name1 (default C)")
    parser.add_argument("--regno-col", default="U", help="Excel column with register number (default AF)")
    parser.add_argument("--sap-supplier-col", default="A", help="Excel column with supplier SAP no. (default A)")
    parser.add_argument("--sap-customer-col", default="B", help="Excel column with customer SAP no. (default B)")
    parser.add_argument("--country-col", default="J", help="Excel column with Country (default J)")

    # excel columns need to be changed, only for change not for extraction
    parser.add_argument("--name1-col", default="T", help="Excel column with Name1 (default T)")
    parser.add_argument("--name2-col", default="D", help="Excel column with Name2 (default D)")
    parser.add_argument("--name3-col", default="E", help="Excel column with Name3 (default E)")
    parser.add_argument("--street-col", default="X", help="Excel column with Street (default X)")
    parser.add_argument("--house-number-col", default="Y", help="Excel column with house number (default Y)")
    parser.add_argument("--city-col", default="Z", help="Excel column with City (default Z)")
    parser.add_argument("--postal-code-col", default="AA", help="Excel column with Postal Code (default AA)") #TODO: check if this is correct
    parser.add_argument("--doc-path-col", default="P", help="Excel column with document stored (default P)")
    parser.add_argument("--changes-check-col", default="Q", help="Excel column with Changes necessary (default Q)")
    parser.add_argument("--date-check-col", default="S", help="Excel column with Date of last check (default S)")
    parser.add_argument("--register-type-col", default="V", help="Excel column with Register type (default AG)")



    parser.add_argument(
        "--start",
        type=int,
        default=3,
        help="Start row (comapnies start at row 3)."
    )
    parser.add_argument(
        "--end",
        type=int,
        default=None,
        help="End row (1-based, inclusive) in the Excel sheet to process (default: last row)."
    )


    return parser


def parse_args(argv: Optional[list[str]] = None):
    return build_parser().parse_args(argv)


def main():
    create_human_check_file()
    args = parse_args()
    try:
        asyncio.run(main_async(args))
    except KeyboardInterrupt:
        print("Aborted by user.")
    except Exception as e:
        print(f"[error] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()



#TODO: Error website ->reload start from the element where it failed, exel empty rows to hold position, over 5 words, umlaute
# python PlayHandelsregister.py -d --download-ad --excel "C:\Users\Nguyen-Bang\Downloads\TestBP.xlsx" --sheet "Tabelle1" --start 25 --end 30
# python PlayHandelsregister.py -s "THYSSENKRUPP SCHULTE GMBH"  --register-number "26718" -d --download-ad
