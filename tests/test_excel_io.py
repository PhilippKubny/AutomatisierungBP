from __future__ import annotations

from collections.abc import Generator
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from bpauto import excel_io
from bpauto.providers.base import CompanyRecord


@pytest.fixture(autouse=True)
def reset_workbook_cache() -> Generator[None, None, None]:
    excel_io.reset()
    yield
    excel_io.reset()


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daten"
    sheet["C2"] = "  "
    sheet["C3"] = "Example GmbH"
    sheet["F3"] = "Musterstraße"
    sheet["G3"] = "1"
    sheet["AA3"] = " 80333 "
    sheet["Z3"] = "München"
    sheet["AC3"] = "DE"
    sheet["C4"] = None
    sheet["C5"] = "Another AG"
    sheet["F5"] = "Beispielallee"
    sheet["G5"] = "2a"
    sheet["AA5"] = "10115"
    sheet["Z5"] = "Berlin"
    workbook.save(path)


def test_iter_rows_trims_and_skips_blank(tmp_path: Path) -> None:
    excel_path = tmp_path / "test.xlsx"
    _create_workbook(excel_path)

    rows = list(
        excel_io.iter_rows(
            excel_path=str(excel_path),
            sheet="Daten",
            start=2,
            end=None,
            name_col="c",
            zip_col="aa",
            city_col="z",
            country_col="ac",
            street_col="f",
            house_number_col="g",
        )
    )

    assert len(rows) == 2
    assert rows[0]["index"] == 3
    assert rows[0]["name"] == "Example GmbH"
    assert rows[0]["zip"] == "80333"
    assert rows[0]["city"] == "München"
    assert rows[0]["country"] == "DE"
    assert rows[0]["street"] == "Musterstraße"
    assert rows[0]["house_number"] == "1"
    assert rows[1]["name"] == "Another AG"


def test_write_and_save(tmp_path: Path) -> None:
    excel_path = tmp_path / "write.xlsx"
    _create_workbook(excel_path)

    record: CompanyRecord = CompanyRecord(
        legal_name="Example GmbH",
        register_type="HRB",
        register_no="12345",
        street="Musterstraße",
        house_number="1",
        zip="80333",
        city="München",
        country="DE",
        notes="confidence=0.90",
        source="northdata_api",
    )

    mapping = {
        "legal_name": "W",
        "register_type": "U",
        "register_no": "V",
        "street": "X",
        "house_number": "Y",
        "zip": "AA",
        "city": "Z",
        "country": "AC",
        "notes": "AB",
        "source": "AD",
    }

    excel_io.write_result(
        excel_path=str(excel_path),
        sheet="Daten",
        row_index=3,
        record=record,
        mapping=mapping,
    )
    excel_io.save(str(excel_path))

    workbook = load_workbook(excel_path)
    sheet = workbook["Daten"]

    assert sheet["W3"].value == "Example GmbH"
    assert sheet["X3"].value == "Musterstraße"
    assert sheet["Y3"].value == "1"
    assert sheet["AA3"].value == "80333"
    assert sheet["AB3"].value == "confidence=0.90"
    assert sheet["AD3"].value == "northdata_api"


def test_missing_sheet_raises(tmp_path: Path) -> None:
    excel_path = tmp_path / "missing.xlsx"
    _create_workbook(excel_path)

    with pytest.raises(ValueError):
        list(
            excel_io.iter_rows(
                excel_path=str(excel_path),
                sheet="Unbekannt",
                start=2,
                end=5,
                name_col="C",
            )
        )
