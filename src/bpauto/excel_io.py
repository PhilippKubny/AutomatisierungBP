"""Hilfsfunktionen für das Lesen und Schreiben von Excel-Arbeitsmappen."""

from __future__ import annotations

from collections.abc import Iterable, Iterator
from datetime import date
from typing import TypedDict, cast

import re

import pandas as pd
from openpyxl import load_workbook

from .providers.base import CompanyRecord
from .utils.logging_setup import setup_logger

_BASE_LOGGER = setup_logger()
LOGGER = _BASE_LOGGER.getChild("excel_io")

try:  # pragma: no cover - ``openpyxl`` always provides these during runtime
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover - fallback für statische Analyse
    Workbook = object  # type: ignore[assignment]
    Worksheet = object  # type: ignore[assignment]


class RowData(TypedDict, total=False):
    """Representation einer gelesenen Tabellenzeile."""

    index: int
    name: str | None
    zip: str | None
    city: str | None
    country: str | None
    street: str | None
    house_number: str | None


_WORKBOOK_CACHE: dict[str, Workbook] = {}

_NAME_NOISE_VALUES = {
    "",
    "-",
    "--",
    "N.N",
    "N.N.",
    "NN",
    "N/A",
    "K.A",
    "K.A.",
    "KA",
    "O.A",
    "O.A.",
    "OA",
    "OHNE ANGABE",
    "KEINE ANGABE",
}

_SUFFIX_SANITISER = re.compile(r"[^0-9A-Z]+")

_VALID_SUFFIX_TOKENS = {
    "AG",
    "AS",
    "AB",
    "BV",
    "BVBA",
    "EG",
    "EK",
    "EV",
    "GBR",
    "GDBR",
    "GMBH",
    "INC",
    "KG",
    "KGA",
    "KGAA",
    "LIMITED",
    "LLC",
    "LLP",
    "LP",
    "LTD",
    "NV",
    "OHG",
    "PARTG",
    "PARTGMBB",
    "PLC",
    "SAS",
    "SARL",
    "SA",
    "SCE",
    "SE",
    "SPA",
    "SRL",
    "STIFTUNG",
    "UG",
    "UGHAFTUNGSBESCHRANKT",
}

_VALID_SUFFIX_PHRASES = {
    "CO KG",
    "CO KGAA",
}


def _get_or_load_workbook(excel_path: str) -> Workbook:
    """Gibt eine zwischengespeicherte Arbeitsmappe zurück."""

    workbook = _WORKBOOK_CACHE.get(excel_path)
    if workbook is None:
        LOGGER.debug("Lade Arbeitsmappe: %s", excel_path)
        workbook = load_workbook(excel_path)
        _WORKBOOK_CACHE[excel_path] = workbook
    return workbook


def _normalise_column(column: str | None) -> str | None:
    if column is None:
        return None
    column = column.strip().upper()
    return column or None


def _cell_to_string(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value_str = value.strip()
    else:
        value_str = str(value).strip()
    return value_str or None


def _is_noise_name_part(value: str) -> bool:
    collapsed = "".join(ch for ch in value.upper() if ch.isalnum())
    normalised = value.strip().upper()
    return collapsed in _NAME_NOISE_VALUES or normalised in _NAME_NOISE_VALUES


def _normalise_name_part(value: str | None) -> str | None:
    if value is None:
        return None
    collapsed = " ".join(value.split())
    if not collapsed:
        return None
    if _is_noise_name_part(collapsed):
        return None
    return collapsed


def _normalise_suffix_token(token: str) -> str:
    return _SUFFIX_SANITISER.sub("", token.upper())


def _has_valid_suffix(name: str | None) -> bool:
    if not name:
        return False
    tokens = [_normalise_suffix_token(part) for part in name.split()]
    tokens = [token for token in tokens if token]
    if not tokens:
        return False
    last_token = tokens[-1]
    if last_token in _VALID_SUFFIX_TOKENS:
        return True
    if len(tokens) >= 2:
        last_two = " ".join(tokens[-2:])
        if last_two in _VALID_SUFFIX_PHRASES:
            return True
        compact_two = "".join(tokens[-2:])
        if compact_two in _VALID_SUFFIX_TOKENS or compact_two in _VALID_SUFFIX_PHRASES:
            return True
    if len(tokens) >= 3:
        last_three = " ".join(tokens[-3:])
        if last_three in _VALID_SUFFIX_PHRASES:
            return True
        compact_three = "".join(tokens[-3:])
        if compact_three in _VALID_SUFFIX_TOKENS or compact_three in _VALID_SUFFIX_PHRASES:
            return True
    return False


def _combine_name_parts(parts: Iterable[str | None]) -> str | None:
    cleaned_parts = [_normalise_name_part(part) for part in parts]
    fragments = [part for part in cleaned_parts if part]
    if not fragments:
        return None
    if len(fragments) == 1:
        return fragments[0]
    combined = " ".join(fragments)
    if _has_valid_suffix(combined):
        return combined
    suffix_fragments = [fragment for fragment in fragments if _has_valid_suffix(fragment)]
    if suffix_fragments:
        return max(suffix_fragments, key=len)
    return max(fragments, key=len)


def _read_cell(worksheet: Worksheet, column: str | None, row_index: int) -> str | None:
    column = _normalise_column(column)
    if not column:
        return None
    cell = worksheet[f"{column}{row_index}"]
    return _cell_to_string(cell.value)


def _find_last_row_with_name(worksheet: Worksheet, name_column: str, start_row: int) -> int:
    normalised_name = _normalise_column(name_column)
    if not normalised_name:
        raise ValueError("Name column must be provided")

    max_row = worksheet.max_row
    for row_idx in range(max_row, start_row - 1, -1):
        value = _read_cell(worksheet, normalised_name, row_idx)
        if value is not None:
            return row_idx
    return start_row - 1


def _get_worksheet(workbook: Workbook, sheet: str | None) -> Worksheet:
    if sheet:
        try:
            return workbook[sheet]
        except KeyError as exc:  # pragma: no cover - defensive
            raise ValueError(f"Arbeitsblatt '{sheet}' wurde nicht gefunden") from exc
    return workbook.active


def _column_letter_to_index(letter: str) -> int:
    normalised = _normalise_column(letter)
    if not normalised:
        raise ValueError("Spaltenbuchstabe darf nicht leer sein")

    idx = 0
    for ch in normalised:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _normalise_sap(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    value_str = str(value).strip()
    return value_str or None


def iter_rows(
    excel_path: str,
    sheet: str,
    start: int,
    end: int | None,
    name_col: str,
    name_additional_cols: tuple[str | None, ...] | None = None,
    zip_col: str | None = None,
    city_col: str | None = None,
    country_col: str | None = None,
    street_col: str | None = None,
    house_number_col: str | None = None,
) -> Iterator[RowData]:
    """Liest Zeilen aus der Arbeitsmappe und liefert bereinigte Werte."""

    workbook = _get_or_load_workbook(excel_path)
    worksheet = _get_worksheet(workbook, sheet)
    normalised_name_col = _normalise_column(name_col)
    if not normalised_name_col:
        raise ValueError("Name column must be provided")

    stop = (
        end if end is not None else _find_last_row_with_name(worksheet, normalised_name_col, start)
    )

    additional_name_cols = tuple(
        col
        for col in (
            _normalise_column(column) for column in (name_additional_cols or ())
        )
        if col
    )

    LOGGER.info("Lese Zeilen %s-%s aus Blatt '%s' (%s)", start, stop, sheet, excel_path)

    def _generator() -> Iterator[RowData]:
        yielded = 0
        if stop < start:
            LOGGER.info("Keine Datenzeilen in Blatt '%s' (%s) gefunden", sheet, excel_path)
            return

        for row_idx in range(start, stop + 1):
            name_value = _read_cell(worksheet, normalised_name_col, row_idx)
            if name_value is None:
                continue

            extra_parts = [
                _read_cell(worksheet, column, row_idx) for column in additional_name_cols
            ]
            combined_name = _combine_name_parts([name_value, *extra_parts])

            row_data: RowData = RowData(
                index=row_idx,
                name=combined_name or name_value,
                zip=_read_cell(worksheet, zip_col, row_idx),
                city=_read_cell(worksheet, city_col, row_idx),
                country=_read_cell(worksheet, country_col, row_idx),
                street=_read_cell(worksheet, street_col, row_idx),
                house_number=_read_cell(worksheet, house_number_col, row_idx),
            )
            yielded += 1
            yield row_data

        LOGGER.info(
            "Verarbeitete Zeilen in Blatt '%s' (%s): %s",
            sheet,
            excel_path,
            yielded,
        )

    return _generator()


def write_result(
    excel_path: str,
    sheet: str,
    row_index: int,
    record: CompanyRecord,
    mapping: dict[str, str],
) -> None:
    """Schreibt Daten aus *record* in die gemappten Spalten."""

    workbook = _get_or_load_workbook(excel_path)
    worksheet = _get_worksheet(workbook, sheet)

    LOGGER.debug(
        "Schreibe Ergebnis für Zeile %s in Blatt '%s' (%s)",
        row_index,
        sheet,
        excel_path,
    )

    record_dict = cast(dict[str, object | None], record)

    for key, column in mapping.items():
        column_letter = _normalise_column(column)
        if not column_letter or key not in record_dict:
            continue
        value = record_dict.get(key)
        if value is None:
            cell_value = ""
        elif isinstance(value, str):
            cell_value = value
        else:
            cell_value = str(value)
        worksheet[f"{column_letter}{row_index}"] = cell_value


def write_to_excel_error(
    path: str,
    sheet: str | None,
    row: int,
    changes_check_col: str,
    error_col: str,
    error_msg: str,
    *,
    pdf_path: str | None = None,
    pdf_path_col: str | None = None,
) -> None:
    """Schreibt Fehlerdetails in die Excel-Datei."""

    workbook = _get_or_load_workbook(path)
    worksheet = _get_worksheet(workbook, sheet)

    worksheet[f"{changes_check_col}{row}"] = "yes"
    worksheet[f"{error_col}{row}"] = error_msg
    if pdf_path and pdf_path_col:
        worksheet[f"{pdf_path_col}{row}"] = pdf_path

    LOGGER.debug("Markiere Zeile %s als fehlerhaft: %s", row, error_msg)


def write_update_to_excel(
    path: str,
    sheet: str | None,
    row: int,
    update_info: dict[str, object],
    *,
    name_col: str,
    regno_col: str,
    sap_supplier_col: str | None,
    sap_customer_col: str | None,
    name2_col: str | None,
    name3_col: str | None,
    street_col: str,
    house_number_col: str,
    city_col: str,
    postal_code_col: str,
    doc_path_col: str,
    changes_check_col: str,
    date_check_col: str,
    register_type_col: str,
    check_file: str | None = None,
) -> bool:
    """Aktualisiert eine Excel-Zeile mit den gelieferten Informationen."""

    df = pd.read_excel(path, sheet_name=sheet, header=None)
    row_index = row - 1

    sup_idx = _column_letter_to_index(sap_supplier_col) if sap_supplier_col else None
    cus_idx = _column_letter_to_index(sap_customer_col) if sap_customer_col else None
    sap_supplier = _normalise_sap(df.iat[row_index, sup_idx]) if sup_idx is not None else None
    sap_customer = _normalise_sap(df.iat[row_index, cus_idx]) if cus_idx is not None else None
    sap_new = _normalise_sap(update_info.get("sap_number"))

    workbook = _get_or_load_workbook(path)
    worksheet = _get_worksheet(workbook, sheet)

    if not sap_new or (sap_new != sap_supplier and sap_new != sap_customer):
        LOGGER.warning(
            "SAP-Mismatch in Zeile %s: neu=%s, Lieferant=%s, Kunde=%s",
            row,
            sap_new,
            sap_supplier,
            sap_customer,
        )
        worksheet[f"{changes_check_col}{row}"] = "yes"
        if check_file:
            with open(check_file, "a", encoding="utf-8") as handle:
                warning_message = (
                    f"[warn] SAP mismatch at row {row}: new={sap_new}, "
                    f"supplier={sap_supplier}, customer={sap_customer}\n"
                )
                handle.write(warning_message)
        return False

    worksheet[f"{name_col}{row}"] = update_info.get("company_name", "")
    if name2_col:
        worksheet[f"{name2_col}{row}"] = update_info.get("company_name_2", "")
    if name3_col:
        worksheet[f"{name3_col}{row}"] = ""

    worksheet[f"{street_col}{row}"] = update_info.get("street", "")
    worksheet[f"{house_number_col}{row}"] = update_info.get("house_number", "")
    worksheet[f"{postal_code_col}{row}"] = update_info.get("postal_code", "")
    worksheet[f"{city_col}{row}"] = update_info.get("city", "")

    worksheet[f"{register_type_col}{row}"] = update_info.get("register_type", "")
    worksheet[f"{regno_col}{row}"] = update_info.get("register_number", "")

    worksheet[f"{doc_path_col}{row}"] = update_info.get("download_path", "")

    worksheet[f"{changes_check_col}{row}"] = "no"
    worksheet[f"{date_check_col}{row}"] = date.today().strftime("%d.%m.%Y")

    LOGGER.debug("Aktualisierte Zeile %s erfolgreich", row)
    return True


def col_letter_to_idx(letter: str) -> int:
    """Konvertiert einen Excel-Spaltenbuchstaben zu einem Null-basierten Index."""

    return _column_letter_to_index(letter)


def read_jobs_from_excel(
    path: str,
    sheet: str | None,
    name_col: str,
    regno_col: str | None,
    sap_supplier_col: str | None,
    sap_customer_col: str | None,
    postal_code_col: str | None,
    country_col: str | None,
    start: int | None,
    end: int | None,
) -> list[dict[str, str | None]]:
    """Liest Unternehmensdaten aus einer Excel-Datei und gibt Jobs zurück."""

    df = pd.read_excel(path, sheet_name=sheet, header=None)
    total_rows = len(df)

    start_row = start if (start and start > 0) else 1
    end_row = end if (end and end > 0) else total_rows

    if start_row > total_rows or end_row < start_row:
        return []

    df_slice = df.iloc[start_row - 1 : end_row]

    def _safe_get(row: pd.Series, column: str | None) -> str | None:
        if not column:
            return None
        idx = _column_letter_to_index(column)
        if idx < 0 or idx >= len(row):
            return None
        value = row.iat[idx]
        if pd.isna(value):
            return None
        if isinstance(value, str):
            cleaned = value.strip()
        else:
            cleaned = str(value).strip()
        return cleaned or None

    jobs: list[dict[str, str | None]] = []
    for _, series in df_slice.iterrows():
        name = _safe_get(series, name_col)
        register_no = _safe_get(series, regno_col)
        sap_supplier = _safe_get(series, sap_supplier_col)
        sap_customer = _safe_get(series, sap_customer_col)
        postal_code = _safe_get(series, postal_code_col)
        country = _safe_get(series, country_col)

        sap_raw = sap_supplier or sap_customer
        if sap_raw is not None and sap_raw.isdigit():
            sap_value: str | None = sap_raw
        else:
            sap_value = sap_raw

        jobs.append(
            {
                "name": name,
                "register_no": register_no,
                "sap": sap_value,
                "postal_code": postal_code,
                "country": country,
            }
        )

    LOGGER.info("Aus Excel geladen: %s Jobs", len(jobs))
    return jobs


def save(excel_path: str) -> None:
    """Persistiert Änderungen auf die Festplatte."""

    workbook = _WORKBOOK_CACHE.get(excel_path)
    if workbook is None:
        LOGGER.debug("Keine Arbeitsmappe im Cache für Pfad: %s", excel_path)
        return
    LOGGER.info("Speichere Arbeitsmappe: %s", excel_path)
    workbook.save(excel_path)


def reset() -> None:
    """Leert den Arbeitsmappen-Cache (hauptsächlich für Tests)."""

    LOGGER.debug("Leere Arbeitsmappen-Cache")
    _WORKBOOK_CACHE.clear()


__all__ = [
    "RowData",
    "iter_rows",
    "write_result",
    "write_to_excel_error",
    "write_update_to_excel",
    "col_letter_to_idx",
    "read_jobs_from_excel",
    "save",
    "reset",
]
