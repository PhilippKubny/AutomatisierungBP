import argparse
import asyncio
import os
import re
import time
import sys
import pandas as pd
from string import ascii_uppercase
from datetime import date
from openpyxl import load_workbook

from playwright.async_api import async_playwright, TimeoutError as PwTimeoutError


def write_to_excel_error(
        path, sheet, row, changes_check_col, error_col, error_msg , pdf_path=None, pdf_path_col=None):
    wb = load_workbook(path)
    ws = wb[sheet] if sheet else wb.active

    ws[f"{changes_check_col}{row}"] = "yes"
    ws[f"{error_col}{row}"] = error_msg
    if pdf_path:
        ws[f"{pdf_path_col}{row}"] = pdf_path
    wb.save(path)
    wb.close()
    return False


def _col_letter_to_idx(letter: str) -> int:
    letter = letter.strip().upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1  # 0-based


def _norm_sap(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def write_update_to_excel(
        path, sheet, row, update_info,
        name_col, regno_col, sap_supplier_col, sap_customer_col,
        name2_col, name3_col, street_col, house_number_col, city_col, postal_code_col,
        doc_path_col, changes_check_col, date_check_col, register_type_col, check_file
):
    """
    HYBRID: Use pandas to check/decide; use openpyxl to write (keeps formatting).
    Returns True if the row was updated.
    """

    # ---- pandas phase: read sheet & check SAP match ----
    df = pd.read_excel(path, sheet_name=sheet, header=None)
    r = row-1

    sup_idx = _col_letter_to_idx(sap_supplier_col) if sap_supplier_col else None
    cus_idx = _col_letter_to_idx(sap_customer_col) if sap_customer_col else None
    sap_supplier = _norm_sap(df.iat[r, sup_idx]) if sup_idx is not None else None
    sap_customer = _norm_sap(df.iat[r, cus_idx]) if cus_idx is not None else None
    sap_new = _norm_sap(update_info.get("sap_number"))

    # ---- openpyxl phase: write only the touched cells (preserve formatting) ----
    wb = load_workbook(path)
    ws = wb[sheet] if sheet else wb.active

    if not sap_new or (sap_new != sap_supplier and sap_new != sap_customer):
        print(f"[warn] SAP mismatch at row {row}: "
              f"new={sap_new}, supplier={sap_supplier}, customer={sap_customer}")
        ws[f"{changes_check_col}{row}"] = "yes"
        f = open(check_file, "a")
        f.write(
            f"[warn] SAP mismatch at row {row}: "
              f"new={sap_new}, supplier={sap_supplier}, customer={sap_customer}")
        return False  # no change

    ws[f"{name_col}{row}"] = update_info.get("company_name", "")
    if name3_col: ws[f"{name3_col}{row}"] = ""  # Name3 blank, as requested

    ws[f"{street_col}{row}"] = update_info.get("street", "")
    ws[f"{house_number_col}{row}"] = update_info.get("house_number", "")
    ws[f"{postal_code_col}{row}"] = update_info.get("postal_code", "")
    ws[f"{city_col}{row}"] = update_info.get("city", "")

    ws[f"{register_type_col}{row}"] = update_info.get("register_type", "")
    ws[f"{regno_col}{row}"] = update_info.get("register_number", "")

    ws[f"{doc_path_col}{row}"] = update_info.get("download_path", "")

    ws[f"{changes_check_col}{row}"] = "no"

    # Write an actual Excel date (keeps number format)
    cell = ws[f"{date_check_col}{row}"]
    cell.value = date.today().strftime("%d-%m-%Y")

    wb.save(path)
    wb.close()
    return True


def col_letter_to_idx(letter: str) -> int:
    """
    Convert an Excel-style column letter into a zero-based index.
    Examples:
        'A'  -> 0
        'B'  -> 1
        ...
        'Z'  -> 25
        'AA' -> 26
    """
    letter = letter.strip().upper()  # Remove whitespace and normalize to uppercase
    idx = 0
    for ch in letter:
        # Shift previous value by 26 and add current letter position
        # 'A' is 1, 'B' is 2, ..., 'Z' is 26
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1  # Convert from 1-based to 0-based index


def read_jobs_from_excel(path, sheet, name_col, regno_col,
                         sap_supplier_col, sap_customer_col, postal_code_col, country_col, start,
                         end):
    """
    Read company data from an Excel file and return a list of jobs to process.

    Parameters:
        path               - Path to the Excel file (e.g., "TestBP.xlsx")
        sheet              - Optional sheet name (None = first sheet)
        name_col           - Column letter for company name (Name1)
        regno_col          - Column letter for register number (AF) or None
        sap_supplier_col   - Column letter for supplier SAP number (A) or None
        sap_customer_col   - Column letter for customer SAP number (other col) or None
        limit              - Maximum number of rows to process (None = all)

    Returns:
        List of dicts with keys: "name", "register_no", "sap"
    """
    # Read the Excel file with header row
    df = pd.read_excel(path, sheet_name=sheet, header=None)  # absolute column positions
    total_rows = len(df)

    # Normalize 1-based bounds → clamp to [1, total_rows]
    s = start if (start and start > 0) else 1
    e = end if (end and end > 0) else total_rows
    if s > total_rows:  # nothing to do
        return []
    if e < s:
        return []

    # Slice rows (convert to 0-based iloc)
    df_slice = df.iloc[s - 1:e]

    def safe_get(row, col_letter: str | None):
        if not col_letter:
            return None
        idx = col_letter_to_idx(col_letter)
        if idx < 0 or idx >= len(row):
            return None
        val = row.iloc[idx]
        return None if (pd.isna(val) or (isinstance(val, str) and val.strip() == "")) else val

    jobs = []
    for _, row in df_slice.iterrows():
        name = safe_get(row, name_col)
        if not name:
            jobs.append({
                "name": None,
                "register_no": None,
                "sap": None,
            })
            continue  # skip rows without a company name

        register_no = safe_get(row, regno_col)
        sap_supplier = safe_get(row, sap_supplier_col)
        sap_customer = safe_get(row, sap_customer_col) if sap_customer_col else None
        postal_code = safe_get(row, postal_code_col)
        country = safe_get(row, country_col)

        # Prefer supplier SAP, else customer SAP, else None
        sap_raw = sap_supplier if sap_supplier not in (None, "") else sap_customer

        # Normalize SAP (Excel often gives numerics as floats)
        if isinstance(sap_raw, float) and sap_raw.is_integer():
            sap = str(int(sap_raw))
        else:
            sap = str(sap_raw).strip() if sap_raw not in (None, "") else None

        # Normalize register number
        if isinstance(register_no, float) and register_no.is_integer():
            register_no = str(int(register_no))
        else:
            register_no = str(register_no).strip() if register_no not in (None, "") else None

        jobs.append({
            "name": str(name).strip(),
            "register_no": register_no,
            "sap": sap,
            "postal_code": str(postal_code).strip() if postal_code not in (None, "") else None,
            "country": str(country).strip() if country not in (None, "") else None,
        })

    return jobs


def main():
    parser = argparse.ArgumentParser(description="Read company data from an Excel file.")

    parser.add_argument("--excel", help="Path to Excel file (e.g., TestBP.xlsx)")
    parser.add_argument("--sheet", default=None, help="Optional sheet name")
    parser.add_argument("--name-col", default="C", help="Excel column with Name1 (default C)")
    parser.add_argument("--regno-col", default="AF", help="Excel column with register number (default AF)")
    parser.add_argument("--sap-supplier-col", default="A", help="Excel column with supplier SAP no. (default A)")
    parser.add_argument("--sap-customer-col", default="B", help="Excel column with customer SAP no. (optional)")
    parser.add_argument(
        "--start",
        type=int,
        default=None,
        help="Start row (1-based) in the Excel sheet to process (default: first row)."
    )
    parser.add_argument(
        "--end",
        type=int,
        default=None,
        help="End row (1-based, inclusive) in the Excel sheet to process (default: last row)."
    )

    args = parser.parse_args()

    jobs = read_jobs_from_excel(
        path=args.excel,
        sheet=args.sheet,
        name_col=args.name_col,
        regno_col=args.regno_col,
        sap_supplier_col=args.sap_supplier_col,
        sap_customer_col=args.sap_customer_col,
        start=args.start,
        end=args.end,
    )
    print(f"[info] loaded {len(jobs)} jobs from Excel (rows {args.start or 1}..{args.end or 'last'})")

    print(f"Found {len(jobs)} jobs:")
    for job in jobs:
        print(job)


def secondary_main():
    """
    This is a secondary main function to demonstrate how to use the write_upadte_to_exel function.
    It can be used for testing or as a standalone script.
    """
    updated = write_update_to_excel(
        path=r"C:\Users\Nguyen-Bang\Downloads\TestBP.xlsx",
        sheet="Tabelle1",
        row=3,
        update_info={
            "sap_number": "11000017",
            "company_name": "ACME GmbH",
            "register_type": "HRB",
            "register_number": "26718",
            "street": "MUSTERSTRASSE",
            "house_number": "12A",
            "postal_code": "80331",
            "city": "MUENCHEN",
            "download_path": r"C:\Users\Nguyen-Bang\Downloads\BP\11000017_CARL KURT WALTHER GMBH & CO. KG_20-08-2025.pdf",
        },
        name_col="C",
        regno_col="AF",
        sap_supplier_col="A",
        sap_customer_col="B",
        name2_col="D",
        name3_col="E",
        street_col="F",
        house_number_col="G",
        city_col="H",
        postal_code_col="I",
        doc_path_col="P",
        changes_check_col="Q",
        date_check_col="S",
        register_type_col="AG",
    )


if __name__ == "__main__":
    secondary_main()
